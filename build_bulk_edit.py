"""
Build the Bulk Edit master workbook (sheet "Appraisal User Qr code Mapping")
from the StockTake inventory file + the QR codes file.

Logic replicates the manual workflow:
  1. Filter StockTake rows where Room = "Inventory".
  2. For each row, XLOOKUP the Location in the QR codes' Description column
     and return the matching UserQrCode.
     Locations annotated like "R58 (AUH-D1-INVOICE-21-04)" are matched
     on the prefix before the first "(".
     OBDT(...) variants all roll up to the OBDT1 QR sticker.
  3. Write two columns: Appraisal code (Deal Id), User Qr Code.
  4. Sort by Deal Id, then Location.

Usage:
  python build_bulk_edit.py
      -> uses the two default input filenames in this folder and writes
         "Bulk Edit - filled.xlsx" next to them.

  python build_bulk_edit.py <inventory.xlsx> <qrcodes.xlsx> [output.xlsx]
"""
import re
import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

HERE = Path(__file__).parent
DEFAULT_INVENTORY = HERE / "INV 21.04.26.xlsx"
DEFAULT_QRCODES = HERE / "QR codes 1.xlsx"

INVENTORY_SHEET = "StockTake Template"
QR_SHEET = "User Qr Code"
OUTPUT_SHEET = "Appraisal User Qr code Mapping"
OUTPUT_HEADERS = ("Appraisal code", "User Qr Code")

ROOM_COL = 0       # column A
LOCATION_COL = 2   # column C
DEAL_ID_COL = 4    # column E

_LOCATION_STRIP_RE = re.compile(r"^([A-Za-z0-9]+)\s*\(")


def build_qr_lookup(qrcodes_source) -> dict:
    """qrcodes_source can be a path or any file-like object openpyxl accepts."""
    wb = load_workbook(qrcodes_source, data_only=True, read_only=True)
    ws = wb[QR_SHEET]
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        qr, desc = row[0], row[1]
        if qr is None or desc is None:
            continue
        key = str(desc).strip()
        lookup.setdefault(key, qr)
    wb.close()
    return lookup


def match_qr(location, lookup: dict):
    if location is None:
        return None
    loc = str(location).strip()
    if loc in lookup:
        return lookup[loc]
    m = _LOCATION_STRIP_RE.match(loc)
    if m:
        prefix = m.group(1)
        if prefix in lookup:
            return lookup[prefix]
        # OBDT rollup: OBDT(<n>) variants share a single QR sticker (OBDT1)
        if prefix.upper() == "OBDT" and "OBDT1" in lookup:
            return lookup["OBDT1"]
    return None


def collect_rows(inventory_source, qr_lookup: dict):
    """inventory_source can be a path or any file-like object.
    Returns (rows, unmatched, skipped_no_deal_id, duplicates).
    duplicates is a list of (deal_id, [(location, qr), ...]) for Deal Ids
    that appear on more than one inventory row.
    """
    wb = load_workbook(inventory_source, data_only=True, read_only=True)
    ws = wb[INVENTORY_SHEET]
    rows = []
    unmatched = []
    skipped_no_deal_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[ROOM_COL] != "Inventory":
            continue
        deal_id = row[DEAL_ID_COL]
        location = row[LOCATION_COL]
        if deal_id is None or str(deal_id).strip().lower() in ("", "no deal id"):
            skipped_no_deal_id += 1
            continue
        qr = match_qr(location, qr_lookup)
        if qr is None:
            unmatched.append((deal_id, location))
        rows.append((deal_id, qr, location))
    wb.close()
    rows.sort(key=lambda r: (str(r[0]), str(r[2]) if r[2] is not None else ""))

    groups = {}
    for deal_id, qr, loc in rows:
        groups.setdefault(deal_id, []).append((loc, qr))
    duplicates = [(d, entries) for d, entries in groups.items() if len(entries) > 1]
    duplicates.sort(key=lambda x: str(x[0]))

    return rows, unmatched, skipped_no_deal_id, duplicates


def write_output(output_target, rows):
    """Build the Bulk Edit master workbook from scratch and save to path or file-like.

    Rows with a Deal Id that appears only once are written first (the "proper"
    block ready for Bulk Edit upload). Rows whose Deal Id appears on multiple
    inventory rows are flagged and appended at the bottom under a marker row,
    with Location included as a third column so the user can reconcile which
    physical unit's QR to keep.
    """
    counts = {}
    for deal_id, _qr, _loc in rows:
        counts[deal_id] = counts.get(deal_id, 0) + 1
    proper = [r for r in rows if counts[r[0]] == 1]
    dupes = [r for r in rows if counts[r[0]] > 1]

    wb = Workbook()
    ws = wb.active
    ws.title = OUTPUT_SHEET
    ws.append(list(OUTPUT_HEADERS))
    for deal_id, qr, _loc in proper:
        ws.append([deal_id, qr])

    if dupes:
        ws.append([])
        marker = ws.cell(row=ws.max_row + 1, column=1,
                         value="REVIEW - DUPLICATE DEAL IDS (pick one per Deal Id, then delete the rest)")
        marker.font = Font(bold=True, color="9C0006")
        marker.fill = PatternFill("solid", fgColor="FFF2CC")
        header_row = ws.max_row + 1
        ws.cell(row=header_row, column=1, value="Appraisal code").font = Font(bold=True)
        ws.cell(row=header_row, column=2, value="User Qr Code").font = Font(bold=True)
        ws.cell(row=header_row, column=3, value="Location").font = Font(bold=True)
        for deal_id, qr, loc in dupes:
            ws.append([deal_id, qr, loc])

    wb.save(output_target)


def main():
    args = sys.argv[1:]
    if len(args) >= 2:
        inventory = Path(args[0])
        qrcodes = Path(args[1])
        output = Path(args[2]) if len(args) >= 3 else HERE / "Bulk Edit - filled.xlsx"
    else:
        inventory = DEFAULT_INVENTORY
        qrcodes = DEFAULT_QRCODES
        output = HERE / "Bulk Edit - filled.xlsx"

    for p in (inventory, qrcodes):
        if not p.exists():
            sys.exit(f"Missing input: {p}")

    print(f"Reading QR codes:  {qrcodes.name}")
    qr_lookup = build_qr_lookup(qrcodes)
    print(f"  {len(qr_lookup)} unique QR descriptions loaded")

    print(f"Reading inventory: {inventory.name}")
    rows, unmatched, skipped, duplicates = collect_rows(inventory, qr_lookup)
    matched = len(rows) - len(unmatched)
    print(f"  {len(rows)} Inventory rows -> matched {matched}, unmatched {len(unmatched)}")
    print(f"  skipped (No Deal Id): {skipped}")
    print(f"  duplicate Deal Ids:   {len(duplicates)}")

    print(f"Writing output:    {output.name}")
    write_output(output, rows)
    print("Done.")

    if unmatched:
        print("\nWARNING: no QR code found for these rows:")
        for deal_id, loc in unmatched[:20]:
            print(f"  {deal_id}  |  location={loc!r}")
        if len(unmatched) > 20:
            print(f"  ... and {len(unmatched) - 20} more")

    if duplicates:
        print("\nWARNING: these Deal Ids appear on multiple inventory rows:")
        for deal_id, entries in duplicates[:20]:
            pairs = ", ".join(f"{loc}->{qr}" for loc, qr in entries)
            print(f"  {deal_id}  |  {pairs}")
        if len(duplicates) > 20:
            print(f"  ... and {len(duplicates) - 20} more")


if __name__ == "__main__":
    main()
